"""
Importador de carga inicial: le a planilha Saldo_Real_RU_*.xlsx e popula o banco
Postgres `ru_custos` (schemas contratos, pagamentos, previsoes, siop).

PRE-REQUISITOS:
  1. Banco criado: `CREATE DATABASE ru_custos;` no Postgres local (porta 5433).
  2. Os servicos contratos-service, pagamentos-service, previsoes-service e siop-service
     ja devem ter subido pelo menos uma vez (mvn spring-boot:run), para o Hibernate
     criar as tabelas de cada schema via ddl-auto=update.
  3. `pip install openpyxl psycopg2-binary`

USO:
  python import_excel.py "caminho/para/Saldo_Real_RU_....xlsx"

O script e idempotente para Campus (nao duplica por nome), mas roda o restante como
INSERT simples — rode uma unica vez sobre um banco vazio. Linhas de rodape ("TOTAL",
"TOTAL GERAL", "Fonte:", notas de metodologia) sao ignoradas.
"""

import os
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
import psycopg2

# Pode ser sobrescrito pela variavel de ambiente RU_CUSTOS_DB_DSN (usado no deploy Docker,
# ex: "host=postgres port=5432 dbname=ru_custos user=postgres password=...").
DB_DSN = os.environ.get(
    "RU_CUSTOS_DB_DSN",
    "host=localhost port=5432 dbname=ru_custos user=postgres password=252107",
)


def parse_valor(v):
    """Converte um valor de celula (numero, string BRL ou vazio) para Decimal ou None."""
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    texto = str(v).strip()
    if not texto or texto in ("—", "-"):
        return None
    if any(c.isalpha() for c in texto):
        return None
    texto = texto.replace(".", "").replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def parse_data(v):
    """Converte celula de data (datetime, ou string dd/mm/aaaa no inicio) para date ou None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})", str(v).strip())
    if m:
        dia, mes, ano = m.groups()
        try:
            return date(int(ano), int(mes), int(dia))
        except ValueError:
            return None
    return None


def linha_valida(valor_primeira_celula):
    if valor_primeira_celula is None:
        return False
    texto = str(valor_primeira_celula).strip().upper()
    if texto in ("", "TOTAL", "TOTAL GERAL"):
        return False
    if texto.startswith("FONTE"):
        return False
    return True


def importar_contratos(wb, conn):
    ws = wb["RU_Contratos"]
    cur = conn.cursor()
    campus_ids = {}
    total_contratos = 0
    total_empenhos = 0

    for row in range(3, ws.max_row + 1):
        campus_nome = ws.cell(row=row, column=1).value
        if not linha_valida(campus_nome):
            continue
        campus_nome = str(campus_nome).strip()

        if campus_nome not in campus_ids:
            cur.execute("SELECT id FROM contratos.campus WHERE nome = %s", (campus_nome,))
            existente = cur.fetchone()
            if existente:
                campus_ids[campus_nome] = existente[0]
            else:
                cur.execute(
                    "INSERT INTO contratos.campus (nome) VALUES (%s) RETURNING id", (campus_nome,)
                )
                campus_ids[campus_nome] = cur.fetchone()[0]
        campus_id = campus_ids[campus_nome]

        empresa = ws.cell(row=row, column=2).value
        numero_contrato = ws.cell(row=row, column=3).value
        processo_contratacao = ws.cell(row=row, column=4).value
        valor_contratual = parse_valor(ws.cell(row=row, column=5).value)
        valor_utilizado = parse_valor(ws.cell(row=row, column=6).value)
        saldo = parse_valor(ws.cell(row=row, column=7).value)
        vigencia_raw = ws.cell(row=row, column=8).value
        vigencia_fim = parse_data(vigencia_raw)
        status = str(vigencia_raw).strip() if vigencia_raw and not vigencia_fim else None

        cur.execute(
            """INSERT INTO contratos.contratos
               (campus_id, empresa, numero_contrato, processo_contratacao,
                valor_contratual, valor_utilizado, saldo, vigencia_fim, status)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (campus_id, empresa, numero_contrato, processo_contratacao,
             valor_contratual, valor_utilizado, saldo, vigencia_fim, status),
        )
        contrato_id = cur.fetchone()[0]
        total_contratos += 1

        numero_ne = ws.cell(row=row, column=9).value
        saldo_empenho = parse_valor(ws.cell(row=row, column=10).value)
        saldo_ne_2025 = parse_valor(ws.cell(row=row, column=13).value)

        if numero_ne or saldo_empenho is not None:
            ano_match = re.match(r"^(\d{4})", str(numero_ne).strip()) if numero_ne else None
            ano = int(ano_match.group(1)) if ano_match else None
            cur.execute(
                """INSERT INTO contratos.empenhos
                   (contrato_id, numero_ne, saldo_empenho, saldo_ne_2025_informativo, ano)
                   VALUES (%s,%s,%s,%s,%s)""",
                (contrato_id, numero_ne, saldo_empenho, saldo_ne_2025, ano),
            )
            total_empenhos += 1

    conn.commit()
    print(f"contratos-service: {len(campus_ids)} campus, {total_contratos} contratos, {total_empenhos} empenhos")


def importar_pagamentos(wb, conn):
    ws = wb[" pagamentos RU 2026"]
    cur = conn.cursor()
    total = 0

    for row in range(3, ws.max_row + 1):
        campus = ws.cell(row=row, column=1).value
        if not linha_valida(campus):
            continue

        mes_referencia = ws.cell(row=row, column=2).value
        empresa = ws.cell(row=row, column=3).value
        numero_contrato = ws.cell(row=row, column=4).value
        numero_processo_contratacao = ws.cell(row=row, column=5).value
        modalidade = ws.cell(row=row, column=6).value
        numero_processo_pagamento = ws.cell(row=row, column=7).value
        recurso = ws.cell(row=row, column=8).value
        ne = ws.cell(row=row, column=9).value
        valor_ne = parse_valor(ws.cell(row=row, column=10).value)
        numero_nf = ws.cell(row=row, column=11).value
        valor_nf = parse_valor(ws.cell(row=row, column=12).value)
        glosa = parse_valor(ws.cell(row=row, column=13).value)
        valor_pago_raw = ws.cell(row=row, column=14).value
        valor_pago = parse_valor(valor_pago_raw)
        observacao = ws.cell(row=row, column=15).value
        if valor_pago is None and valor_pago_raw and not observacao:
            observacao = str(valor_pago_raw).strip()

        cur.execute(
            """INSERT INTO pagamentos.pagamentos
               (campus, mes_referencia, empresa, numero_contrato, numero_processo_contratacao,
                modalidade, numero_processo_pagamento, recurso, ne, valor_ne, numero_nf,
                valor_nf, glosa, valor_pago, observacao, ano)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (campus, mes_referencia, empresa, numero_contrato, numero_processo_contratacao,
             modalidade, numero_processo_pagamento, recurso, ne, valor_ne, numero_nf,
             valor_nf, glosa, valor_pago, observacao, 2026),
        )
        total += 1

    conn.commit()
    print(f"pagamentos-service: {total} pagamentos")


MESES_EXECUCAO = [
    (10, date(2025, 12, 1), "EXECUTADO"),
    (11, date(2026, 1, 1), "EXECUTADO"),
    (12, date(2026, 2, 1), "EXECUTADO"),
    (13, date(2026, 3, 1), "EXECUTADO"),
    (14, date(2026, 4, 1), "EXECUTADO"),
    (15, date(2026, 5, 1), "EXECUTADO"),
    (16, date(2026, 6, 1), "EXECUTADO"),
    (17, date(2026, 7, 1), "EXECUTADO"),
    (18, date(2026, 8, 1), "PREVISAO"),
    (19, date(2026, 9, 1), "PREVISAO"),
    (20, date(2026, 10, 1), "PREVISAO"),
    (21, date(2026, 11, 1), "PREVISAO"),
    (22, date(2026, 12, 1), "PREVISAO"),
    (23, date(2027, 1, 1), "PREVISAO"),
]


def importar_previsoes(wb, conn):
    ws = wb["Previsões 2026"]
    cur = conn.cursor()
    total_contratos = 0
    total_execucoes = 0

    for row in range(14, 28):
        favorecido = ws.cell(row=row, column=3).value
        if not linha_valida(favorecido) or (favorecido and len(str(favorecido)) > 200):
            continue

        numero_contrato = ws.cell(row=row, column=4).value
        vigencia_inicio = parse_data(ws.cell(row=row, column=5).value)
        vigencia_fim = parse_data(ws.cell(row=row, column=6).value)
        plano_interno = ws.cell(row=row, column=7).value
        setor_sipac = ws.cell(row=row, column=8).value
        objeto = ws.cell(row=row, column=9).value
        valor_contrato = parse_valor(ws.cell(row=row, column=30).value) or Decimal("0")
        empenhado = parse_valor(ws.cell(row=row, column=25).value) or Decimal("0")
        numero_ne = ws.cell(row=row, column=27).value
        processo_sei = ws.cell(row=row, column=35).value
        sobra_aproveitavel = str(ws.cell(row=row, column=48).value or "").strip().upper() == "S"

        cur.execute(
            """INSERT INTO previsoes.contratos_previsao
               (favorecido, numero_contrato, vigencia_inicio, vigencia_fim, plano_interno,
                setor_sipac, objeto, valor_contrato, numero_ne, empenhado, processo_sei,
                sobra_aproveitavel)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (favorecido, numero_contrato, vigencia_inicio, vigencia_fim, plano_interno,
             setor_sipac, objeto, valor_contrato, numero_ne, empenhado, processo_sei,
             sobra_aproveitavel),
        )
        contrato_previsao_id = cur.fetchone()[0]
        total_contratos += 1

        for coluna, mes, tipo in MESES_EXECUCAO:
            valor = parse_valor(ws.cell(row=row, column=coluna).value)
            if valor is None:
                continue
            cur.execute(
                """INSERT INTO previsoes.execucoes_mensais
                   (contrato_previsao_id, mes_referencia, tipo, valor)
                   VALUES (%s,%s,%s,%s)""",
                (contrato_previsao_id, mes, tipo, valor),
            )
            total_execucoes += 1

    conn.commit()
    print(f"previsoes-service: {total_contratos} contratos, {total_execucoes} execucoes mensais")


def importar_siop(wb, conn):
    ws = wb["SIOP"]
    cur = conn.cursor()
    total_planos = 0
    total_acoes = 0

    for row in range(3, 7):
        plano = ws.cell(row=row, column=2).value
        if not linha_valida(plano):
            continue
        localizador = ws.cell(row=row, column=1).value
        projeto_lei = parse_valor(ws.cell(row=row, column=3).value) or Decimal("0")
        dotacao_inicial = parse_valor(ws.cell(row=row, column=4).value) or Decimal("0")
        dotacao_atual = parse_valor(ws.cell(row=row, column=5).value) or Decimal("0")
        empenhado = parse_valor(ws.cell(row=row, column=6).value) or Decimal("0")
        liquidado = parse_valor(ws.cell(row=row, column=7).value) or Decimal("0")
        pago = parse_valor(ws.cell(row=row, column=8).value) or Decimal("0")

        cur.execute(
            """INSERT INTO siop.planos_orcamentarios
               (localizador, plano_orcamentario, projeto_lei, dotacao_inicial, dotacao_atual,
                empenhado, liquidado, pago)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
            (localizador, plano, projeto_lei, dotacao_inicial, dotacao_atual, empenhado, liquidado, pago),
        )
        total_planos += 1

    def importar_acoes(linhas, tipo):
        nonlocal total_acoes
        for row in linhas:
            discriminacao = ws.cell(row=row, column=2).value
            if not linha_valida(discriminacao):
                continue
            acao = ws.cell(row=row, column=1).value or "4002 - Assistência ao Estudante de Ensino Superior"
            valores = [parse_valor(ws.cell(row=row, column=c).value) or Decimal("0") for c in range(3, 13)]
            cur.execute(
                """INSERT INTO siop.acoes_orcamentarias
                   (acao, discriminacao, tipo, dotacao_ploa_custeio, dotacao_anulada_ploa,
                    dotacao_inicial_loa_custeio, recomposicao_ploa, dotacao_autorizada,
                    remanejado_cancelado, credito_suplementar, dotacao_atualizada, executado,
                    saldo_atualizar)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (acao, discriminacao, tipo, *valores),
            )
            total_acoes += 1

    importar_acoes(range(14, 18), "CUSTEIO")
    importar_acoes(range(26, 27), "CAPITAL")

    conn.commit()
    print(f"siop-service: {total_planos} planos orcamentarios, {total_acoes} acoes orcamentarias")


def main():
    if len(sys.argv) != 2:
        print("uso: python import_excel.py <caminho_para_o_xlsx>")
        sys.exit(1)

    caminho = sys.argv[1]
    print(f"lendo {caminho}...")
    wb = openpyxl.load_workbook(caminho, data_only=True)

    conn = psycopg2.connect(DB_DSN)
    try:
        importar_contratos(wb, conn)
        importar_pagamentos(wb, conn)
        importar_previsoes(wb, conn)
        importar_siop(wb, conn)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    print("importacao concluida.")


if __name__ == "__main__":
    main()
